import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.repositories.estudiante_repository import EstudianteRepository

estudiante_repo = EstudianteRepository()

RISK_COLORS = {
    "verde":    "FF22C55E",
    "amarillo": "FFEAB308",
    "rojo":     "FFEF4444",
}


class ReportService:
    def generate_excel(self, maestro_id: int = None) -> bytes:
        estudiantes = estudiante_repo.get_filtered(maestro_id=maestro_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes RCM"

        headers = ["ID", "Nombre", "Grupo", "Participación %", "Nivel de Riesgo"]
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill(start_color="FF1A365D", end_color="FF1A365D", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, est in enumerate(estudiantes, 2):
            d = est.to_dict()
            ws.cell(row=row_idx, column=1, value=est.id)
            ws.cell(row=row_idx, column=2, value=est.nombre)
            ws.cell(row=row_idx, column=3, value=est.grupo)
            ws.cell(row=row_idx, column=4, value=d["participacion"])

            risk_cell = ws.cell(row=row_idx, column=5, value=est.nivel_riesgo.upper())
            color = RISK_COLORS.get(est.nivel_riesgo, "FFCCCCCC")
            risk_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            risk_cell.font = Font(
                bold=True,
                color="FFFFFFFF" if est.nivel_riesgo != "amarillo" else "FF000000"
            )
            risk_cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def generate_pdf_html(self, maestro_id: int = None) -> str:
        estudiantes = estudiante_repo.get_filtered(maestro_id=maestro_id)
        stats = estudiante_repo.get_stats()

        rows = ""
        for est in estudiantes:
            d = est.to_dict()
            color_map = {"verde": "#22C55E", "amarillo": "#EAB308", "rojo": "#EF4444"}
            color = color_map.get(est.nivel_riesgo, "#ccc")
            rows += f"""
            <tr>
              <td>{est.id}</td>
              <td>{est.nombre}</td>
              <td>{est.grupo or '—'}</td>
              <td>{d['participacion']}%</td>
              <td style="color:{color}; font-weight:bold;">{est.nivel_riesgo.upper()}</td>
            </tr>"""

        return f"""<!DOCTYPE html>
        <html>
        <head>
          <meta charset="UTF-8">
          <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; color: #1a1a1a; }}
            h1 {{ color: #1A365D; border-bottom: 2px solid #1A365D; padding-bottom: 8px; }}
            .stats {{ display: flex; gap: 16px; margin: 20px 0; }}
            .stat {{ background: #f1f5f9; padding: 12px 20px; border-radius: 8px; text-align: center; }}
            .stat strong {{ display: block; font-size: 24px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #1A365D; color: white; padding: 10px; text-align: left; }}
            td {{ padding: 8px 10px; border-bottom: 1px solid #e2e8f0; }}
            tr:nth-child(even) {{ background: #f8fafc; }}
          </style>
        </head>
        <body>
          <h1>Reporte RCM — Estudiantes</h1>
          <div class="stats">
            <div class="stat"><strong>{stats['total']}</strong>Total</div>
            <div class="stat" style="border-left:4px solid #22C55E"><strong>{stats['verde']}</strong>Sin Riesgo</div>
            <div class="stat" style="border-left:4px solid #EAB308"><strong>{stats['amarillo']}</strong>Riesgo Medio</div>
            <div class="stat" style="border-left:4px solid #EF4444"><strong>{stats['rojo']}</strong>Riesgo Crítico</div>
          </div>
          <table>
            <thead>
              <tr><th>ID</th><th>Nombre</th><th>Grupo</th><th>Participación</th><th>Riesgo</th></tr>
            </thead>
            <tbody>{rows}</tbody>
          </table>
        </body>
        </html>"""
